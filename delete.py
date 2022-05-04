def dar(sr,sc,fnum1,fnum2):
    import openpyxl
    import pandas
    import numpy
    #import xlrd

   # str1 = '\''
    #input_file = str(str1 + input1 + str1)
    #Open workbook Sheet1
    wb = openpyxl.load_workbook(sr)
    ws = wb['Sheet1']
    #Delete
    for rowNum in range(1,ws.max_row+1):
        for columnNum in range (1,ws.max_column+1):
            data1 = ws.cell(row=rowNum,column=columnNum).value
            if '(跳过)' in str(data1):
                ws.cell(row=rowNum,column=columnNum).value = ''


#    data2 = pandas.read_excel(sc,sheet_name='Sheet1')
#    data2 = data2.applymap(lambda x:'{:.2f}'.format(x)).applymap(lambda x:eval(x))
    for rowNum in range(2,ws.max_row+1):
        for columnNum in range (fnum1,fnum2+1):
            num1 =ws.cell(row=rowNum,column=columnNum).value
            num2 = float(num1)
            ws.cell(row=rowNum,column=columnNum).value = num2
    # Save File
    wb.save(sc)

# file_sr,file_sc = input().split(",")
#file_sr = '1.xlsx'
#file_sc = 'calculate1.xlsx'
#dar(file_sr,file_sc)

