import openpyxl
#coding = utf -8
import pandas
import os
#引入dar
from delete import dar
from realjudge import judge
from loaddir import movefile
import tkinter as tk
from tkinter import filedialog


# 使用dar删除跳过
# print('input file_sr,file_sc')
# file_sr,file_sc = input().split(",")
# dar(file_sr,file_sc)
print('请把问卷导出的Excel表格拖入程序内并按回车键，如有问题可联系作者QQ：1793703473')
print('input file_sr')

file_sr = input()#发布用
# file_sr = 'E:/OneDrive-E5/OneDrive - 7cvmmb/404/python/xidajiaoxue/3.xlsx'#调试用
dirname, filename = os.path.split(file_sr)
print(dirname)
print(filename)

newdir = movefile(filename)
print(newdir)

file_sr = filename
file_sc = "Calculate_"+str(file_sr)
#命名新输出文件
file_sc2 = "Calculate2_"+str(file_sr)
file_sc3 = "summary_"+str(file_sr)
#得出总列数
jnum = judge(file_sr)
print('judge:'+str(jnum))
#计算分数所在列
fnum1=jnum-17
fnum2=jnum-2
#计算总评所在列
zhnum=jnum-1
#计算老师最后一列
lsnum=jnum-18
#计算文字评价所在列
pj=jnum
print('lsnum:'+str(lsnum))
dar(file_sr,file_sc,fnum1,fnum2)

#file_sr =input()
# file_sr ='1.xlsx'
# file_sc = 'calculate1.xlsx'
# dar(file_sr,file_sc)

#打开calculate1中的活跃Sheet
#print('Open file_sc')
wb2 = openpyxl.load_workbook(file_sc)
sheet2 = wb2.active
#新建excel文件，并打开活跃Sheet
wb3 = openpyxl.Workbook()
sheet3 = wb3.active

#将课程和老师一列写入新文件
if lsnum == 7:
    #遍历第二行-最后一行
    for rowNum in range(2, sheet2.max_row+1):
        #设置每一行初始分数i=0
        i=str()
        #遍历第7列，并把每个单元格的数值加到i上
        #for columnNum in range (7):
        i=sheet2.cell(row=rowNum,column=7).value
        #将i写入所在行的第30列单元格
        #sheet2.cell(row=rowNum, column=30).value = i
        sheet3.cell(row=rowNum, column=1).value = i
if lsnum != 7:
    # 将老师所在几列数据合并写在judge+2列
    # 遍历第二行-最后一行
    for rowNum in range(2, sheet2.max_row + 1):
        # 设置每一行初始老师名称j为空
        j = str()
        # 遍历第8列-lsnum列，并把每个单元格的文本合并到j上
        for columnNum in range(8, lsnum+1):
            if sheet2.cell(row=rowNum, column=columnNum).value != None:
                j = str(j) + str(sheet2.cell(row=rowNum, column=columnNum).value)
        sheet2.cell(row=rowNum, column=jnum+2).value = j
    # 将课程-老师一列写入新文件
    # 遍历第二行-最后一行
    for rowNum in range(2, sheet2.max_row + 1):
        # 设置每一行初始分数i=0
        i = str()
        # 遍历第7列，并把每个单元格的数值加到i上
        # for columnNum in range (7):
        i = str(sheet2.cell(row=rowNum, column=7).value) + str('-') + str(sheet2.cell(row=rowNum, column=jnum+2).value)
        # 将i写入所在行的第30列单元格
        # sheet2.cell(row=rowNum, column=30).value = i
        sheet3.cell(row=rowNum, column=1).value = i

#求和(fnum1,fnum2)，并写入第jnum+1列
#遍历第二行-最后一行
for rowNum in range(2, sheet2.max_row+1):
    #设置每一行初始分数i=0
    i=0
    #遍历第8列-25列，并把每个单元格的数值加到i上
    for columnNum in range (fnum1,fnum2+1):
        i=i+sheet2.cell(row=rowNum,column=columnNum).value
    #将i写入所在行的第30列单元格
    sheet2.cell(row=rowNum, column=jnum+1).value = i
    #sheet2.cell(row=rowNum, column=2).value = i

#将分数一列写入新文件
#遍历第二行-最后一行
for rowNum in range(2, sheet2.max_row+1):
    #设置每一行初始分数i=0
    i=0
    #遍历第7列，并把每个单元格的数值加到i上
    #for columnNum in range (7):
    i=sheet2.cell(row=rowNum,column=jnum+1).value
    #将i写入所在行的第30列单元格
    #sheet2.cell(row=rowNum, column=30).value = i
    sheet3.cell(row=rowNum, column=2).value = i

#将总体评价一列写入新文件
#遍历第二行-最后一行
for rowNum in range(2, sheet2.max_row+1):
    #设置每一行初始分数i=0
    i=0
    #遍历第7列，并把每个单元格的数值加到i上
    #for columnNum in range (7):
    i=sheet2.cell(row=rowNum,column=zhnum).value
    #将i写入所在行的第30列单元格
    #sheet2.cell(row=rowNum, column=30).value = i
    sheet3.cell(row=rowNum, column=4).value = float(i)

#将评价一列写入新文件
#遍历第二行-最后一行
for rowNum in range(2, sheet2.max_row+1):
    #设置每一行初始分数i=0
    i=str()
    #遍历第7列，并把每个单元格的数值加到i上
    #for columnNum in range (7):
    i=sheet2.cell(row=rowNum,column=pj).value
    #将i写入所在行的第30列单元格
    #sheet2.cell(row=rowNum, column=30).value = i
    sheet3.cell(row=rowNum, column=3).value = i


sheet3.cell(row=1, column=1).value = str('课程-老师')
sheet3.cell(row=1, column=2).value = str('总分')
sheet3.cell(row=1, column=3).value = str('意见和建议')
sheet3.cell(row=1, column=4).value = str('总体评价')

wb3.save(str(dirname)+'\\'+str(newdir)+'\\'+file_sc2)
print(str(dirname)+'\\'+str(newdir)+'\\'+file_sc2)

print('Saving file_sc')
wb2.save(str(dirname)+'\\'+str(newdir)+'\\'+file_sc)
print('Calculate Done')

#建立字典
jxpj = {}
jxpj2 = {}
#打开文件
wb4 = openpyxl.load_workbook(str(dirname)+'\\'+str(newdir)+'\\'+file_sc2)
ws4 = wb4.active
for row in range(2, ws4.max_row+1):
    #从calculate1.xlsx中获取课程名称，老师名称，总分
    kcmc = ws4['A' + str(row)].value
    zf = ws4['B' + str(row)].value
    yj = ws4['C' + str(row)].value
    ztpj = ws4['D' + str(row)].value
    jxpj.setdefault(kcmc,{'zrs':0,'zf':0,'ztpj':0})
    jxpj[kcmc]['zrs'] +=1
    jxpj[kcmc]['zf'] += float(zf)
    jxpj[kcmc]['ztpj'] += float(ztpj)
    #从calculate1.xlsx中获取获取评价
    jxpj2.setdefault(kcmc,{})
    jxpj2[kcmc].setdefault(yj,{})
#print(jxpj)
#print(jxpj2)
#将数据写入excel文件
wb5 = openpyxl.Workbook()
ws5 = wb5.active
numls = 2
for wls in jxpj:
    ws5.cell(row=numls, column=1).value = wls
    ws5.cell(row=numls, column=2).value = jxpj[wls]['zrs']
    ws5.cell(row=numls, column=3).value = jxpj[wls]['zf']
    ws5.cell(row=numls, column=5).value = jxpj[wls]['ztpj']
    #colnum1=7
    #for lsyj in jxpj2[wls]:
    #    ws5.cell(row=numls, column=colnum1).value = lsyj
    #    colnum1 +=1
    numls +=1


#for row in range(2, ws4.max_row+1):
#    lsc=ws4.cell(row=row, column=1).value
#    lsyj=ws4.cell(row=row, column=3).value

#    for i in range(2, ws5.max_row+1):
#        j=7
#        if lsc == ws5.cell(row=i, column=1).value:
#            ws5.cell(row=i, column=j).value = lsyj
#            j +=1
print(ws5.max_row+1)
for row in range(2, ws5.max_row+1):
    lsc=ws5.cell(row=row, column=1).value
    j=7
    for i in range(2, ws4.max_row+1):
        if lsc == ws4.cell(row=i, column=1).value:
            lsyj=ws4.cell(row=i, column=3).value
            ws5.cell(row=row, column=j).value = lsyj
            j +=1
    ws5.cell(row=row, column=4).value = (ws5.cell(row=row, column=3).value)/(ws5.cell(row=row, column=2).value)
    ws5.cell(row=row, column=6).value = (ws5.cell(row=row, column=5).value)/(ws5.cell(row=row, column=2).value)
ws5.cell(row=1, column=1).value = str('课程-老师')
ws5.cell(row=1, column=2).value = str('填写人数')
ws5.cell(row=1, column=3).value = str('总分')
ws5.cell(row=1, column=4).value = str('平均分')
ws5.cell(row=1, column=5).value = str('总体评价总分')
ws5.cell(row=1, column=6).value = str('总体评价平均分')
ws5.cell(row=1, column=7).value = str('意见和建议')
wb5.save(str(dirname)+'\\'+str(newdir)+'\\'+file_sc3)
#删除多余文件
os.remove(file_sc)